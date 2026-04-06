"""
modules/exporter.py
===================
Экспорт результатов парсинга в Excel и CSV.

Excel-отчёт содержит несколько листов:
  1. Все ключи (Wordstat + Serpstat) с частотностями и спамностью
  2. Очищенные ключи после стоп-слов
  3. Сезонность (Google Trends)
  4. Кластеры LSI со структурой и интентами
  5. Анализ конкурентов
  6. Сводная статистика

Стили: условное форматирование, автоширина колонок, заморозка заголовков.
"""

import csv
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional

from loguru import logger

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl не установлен — Excel-экспорт недоступен")


# ─── Цветовая схема ───────────────────────────────────────────────────────────

COLORS = {
    "header_bg":    "1F3864",   # тёмно-синий
    "header_font":  "FFFFFF",
    "subheader_bg": "2E75B6",
    "subheader_font": "FFFFFF",
    "row_even":     "DCE6F1",
    "row_odd":      "FFFFFF",
    "green":        "70AD47",
    "yellow":       "FFD966",
    "red":          "FF0000",
    "orange":       "F4B942",
    "light_green":  "E2EFDA",
    "light_red":    "FFDAD9",
}


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> "Font":
    return Font(bold=bold, color=color, size=size)


def _border() -> "Border":
    thin = Side(style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ─── Основной класс ───────────────────────────────────────────────────────────

class KeywordExporter:
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ── Excel ─────────────────────────────────────────────────────────────────

    def to_excel(
        self,
        wordstat_results: list,
        keywords_clean: List[str],
        seasonal_data: Dict[str, dict],
        clusters: Dict[str, dict],
        competitor_data: Optional[Dict] = None,
        filename: Optional[str] = None,
    ) -> str:
        """
        Создаёт многостраничный Excel-отчёт.

        Returns:
            Путь к созданному файлу
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("Пропускаем Excel-экспорт (нет openpyxl)")
            return ""

        if filename is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            filename = f"seo_keywords_{ts}.xlsx"

        path = self.output_dir / filename
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # удаляем дефолтный лист

        # Листы
        self._sheet_all_keywords(wb, wordstat_results)
        self._sheet_clean_keywords(wb, keywords_clean, seasonal_data)
        self._sheet_seasonal(wb, seasonal_data)
        self._sheet_clusters(wb, clusters)
        if competitor_data:
            self._sheet_competitors(wb, competitor_data)
        self._sheet_summary(wb, wordstat_results, keywords_clean, clusters, seasonal_data)

        wb.save(path)
        logger.success(f"Excel сохранён: {path}")
        return str(path)

    def _sheet_all_keywords(self, wb, wordstat_results: list):
        ws = wb.create_sheet("Все ключи (Wordstat)")
        headers = [
            "Ключевое слово", "Базовая частота", "Точная [!]",
            "Фразовая \"\"", "Спамность", "Регион", "Источник"
        ]
        self._write_headers(ws, headers)

        for i, r in enumerate(wordstat_results, start=2):
            d = r.to_dict() if hasattr(r, "to_dict") else r
            row = [
                d.get("keyword", ""),
                d.get("shows_base", 0),
                d.get("shows_exact", ""),
                d.get("shows_phrase", ""),
                d.get("spaminess", ""),
                d.get("region_id", ""),
                d.get("source", "wordstat"),
            ]
            ws.append(row)
            fill = _fill(COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"])
            for cell in ws[i]:
                cell.fill = fill
                cell.border = _border()
                cell.alignment = Alignment(vertical="center")

        # Условное форматирование: спамность (колонка E)
        if len(wordstat_results) > 0:
            ws.conditional_formatting.add(
                f"E2:E{len(wordstat_results)+1}",
                ColorScaleRule(
                    start_type="min", start_color="70AD47",
                    mid_type="num", mid_value=5, mid_color="FFD966",
                    end_type="max", end_color="FF0000"
                )
            )

        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_clean_keywords(self, ws_parent, keywords: List[str], seasonal: dict):
        ws = ws_parent.create_sheet("Очищенные ключи")
        headers = ["Ключевое слово", "Слов в запросе", "Сезонный?", "Пик спроса", "Коэффициент сезонности"]
        self._write_headers(ws, headers)

        for i, kw in enumerate(keywords, start=2):
            s = seasonal.get(kw, {})
            row = [
                kw,
                len(kw.split()),
                "Да" if s.get("is_seasonal") else "Нет",
                s.get("peak_month", ""),
                s.get("ratio", ""),
            ]
            ws.append(row)
            fill = _fill(COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"])
            for cell in ws[i]:
                cell.fill = fill
                cell.border = _border()

        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_seasonal(self, wb, seasonal_data: dict):
        ws = wb.create_sheet("Сезонность (Trends)")
        headers = [
            "Ключевое слово", "Тип сезонности", "Пиковый месяц",
            "Минимальный месяц", "Коэф. сезонности",
            "Янв", "Фев", "Мар", "Апр", "Май", "Июн",
            "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"
        ]
        self._write_headers(ws, headers)

        for i, (kw, data) in enumerate(seasonal_data.items(), start=2):
            monthly = data.get("monthly_avg", {})
            row = [
                kw,
                data.get("label", ""),
                data.get("peak_month", ""),
                data.get("low_month", ""),
                data.get("ratio", ""),
                *[monthly.get(m, "") for m in range(1, 13)]
            ]
            ws.append(row)
            fill = _fill(COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"])
            for cell in ws[i]:
                cell.fill = fill
                cell.border = _border()

        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_clusters(self, wb, clusters: dict):
        ws = wb.create_sheet("LSI Кластеры")
        headers = [
            "Кластер (главный ключ)", "Интент", "Размер",
            "Все ключи кластера"
        ]
        self._write_headers(ws, headers)

        # Цвета по интентам
        INTENT_COLORS = {
            "транзакционный": "C6EFCE",
            "информационный": "DDEBF7",
            "навигационный":  "FFF2CC",
            "коммерческий":   "FCE4D6",
            "смешанный":      "EDEDED",
        }

        for i, (name, data) in enumerate(sorted(
            clusters.items(), key=lambda x: x[1].get("size", 0), reverse=True
        ), start=2):
            intent = data.get("intent", "смешанный")
            row = [
                data.get("main_keyword", name),
                intent,
                data.get("size", len(data.get("keywords", []))),
                " | ".join(data.get("keywords", [])),
            ]
            ws.append(row)
            color = INTENT_COLORS.get(intent, "FFFFFF")
            fill = _fill(color)
            for cell in ws[i]:
                cell.fill = fill
                cell.border = _border()
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 80
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20

    def _sheet_competitors(self, wb, competitor_data: dict):
        ws = wb.create_sheet("Анализ конкурентов")
        headers = [
            "Запрос", "URL", "Домен", "Title", "H1",
            "Слов", "Изображений", "Schema.org", "FAQ", "Хлебные крошки",
            "CMS", "Время загрузки (мс)"
        ]
        self._write_headers(ws, headers)

        row_n = 2
        for query, data in competitor_data.items():
            pages = data.get("pages", []) if isinstance(data, dict) else data
            for page in pages:
                if isinstance(page, dict):
                    row = [
                        query,
                        page.get("url", ""),
                        page.get("domain", ""),
                        page.get("title", ""),
                        page.get("h1", ""),
                        page.get("word_count", ""),
                        page.get("images_count", ""),
                        page.get("schema_types", ""),
                        "Да" if page.get("has_faq") else "Нет",
                        "Да" if page.get("has_breadcrumbs") else "Нет",
                        page.get("cms", ""),
                        page.get("load_time_ms", ""),
                    ]
                    ws.append(row)
                    fill = _fill(COLORS["row_even"] if row_n % 2 == 0 else COLORS["row_odd"])
                    for cell in ws[row_n]:
                        cell.fill = fill
                        cell.border = _border()
                    row_n += 1

        self._auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_summary(self, wb, wordstat, clean, clusters, seasonal):
        ws = wb.create_sheet("Сводка", 0)  # первый лист
        ws.sheet_view.showGridLines = False

        # Заголовок
        ws["A1"] = "SEO Keyword Parser — Отчёт"
        ws["A1"].font = Font(bold=True, size=16, color=COLORS["header_bg"])
        ws["A2"] = f"Сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws["A2"].font = Font(color="666666", size=10)
        ws.append([])

        stats = [
            ("Всего ключей (Wordstat)", len(wordstat)),
            ("После фильтрации", len(clean)),
            ("Сезонных ключей", sum(1 for d in seasonal.values() if d.get("is_seasonal"))),
            ("LSI-кластеров", len(clusters)),
            ("Транзакционных кластеров",
             sum(1 for d in clusters.values() if d.get("intent") == "транзакционный")),
            ("Информационных кластеров",
             sum(1 for d in clusters.values() if d.get("intent") == "информационный")),
        ]

        ws.append(["Показатель", "Значение"])
        header_row = ws.max_row
        ws[f"A{header_row}"].fill = _fill(COLORS["header_bg"])
        ws[f"A{header_row}"].font = _font(bold=True, color="FFFFFF")
        ws[f"B{header_row}"].fill = _fill(COLORS["header_bg"])
        ws[f"B{header_row}"].font = _font(bold=True, color="FFFFFF")

        for label, value in stats:
            ws.append([label, value])
            r = ws.max_row
            ws[f"A{r}"].border = _border()
            ws[f"B{r}"].border = _border()
            ws[f"B{r}"].font = Font(bold=True, size=11)

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15

    # ── CSV ───────────────────────────────────────────────────────────────────

    def to_csv(
        self,
        keywords: List[str],
        seasonal_data: Optional[Dict] = None,
        filename: Optional[str] = None,
    ) -> str:
        """Экспортирует очищенные ключи в CSV."""
        if filename is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            filename = f"keywords_clean_{ts}.csv"

        path = self.output_dir / filename
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["keyword", "word_count", "seasonal", "peak_month", "ratio"])
            for kw in keywords:
                s = (seasonal_data or {}).get(kw, {})
                writer.writerow([
                    kw,
                    len(kw.split()),
                    "1" if s.get("is_seasonal") else "0",
                    s.get("peak_month", ""),
                    s.get("ratio", ""),
                ])

        logger.success(f"CSV сохранён: {path}")
        return str(path)

    def clusters_to_csv(self, clusters: Dict[str, dict], filename: Optional[str] = None) -> str:
        """Экспортирует кластеры в CSV."""
        if filename is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            filename = f"clusters_{ts}.csv"

        path = self.output_dir / filename
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["cluster", "intent", "size", "keyword"])
            for cluster_name, data in clusters.items():
                for kw in data.get("keywords", []):
                    writer.writerow([
                        data.get("main_keyword", cluster_name),
                        data.get("intent", ""),
                        data.get("size", ""),
                        kw,
                    ])

        logger.success(f"Кластеры CSV: {path}")
        return str(path)

    # ── Хелперы ───────────────────────────────────────────────────────────────

    def _write_headers(self, ws, headers: List[str]):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = _fill(COLORS["header_bg"])
            cell.font = _font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _border()
        ws.row_dimensions[1].height = 25

    def _auto_width(self, ws, max_width: int = 60):
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max(
                (len(str(cell.value or "")) for cell in col),
                default=10
            )
            ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)
